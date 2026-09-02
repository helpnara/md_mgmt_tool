"""첨부 파일 저장.

용량 상한이 없으므로 전체를 메모리에 올리지 않고 청크 단위로 흘려보낸다.
저장 위치는 과제 폴더 안의 상대 경로라, 외부 뷰어에서도 링크가 그대로 열린다.
"""
from __future__ import annotations

import hashlib
import mimetypes
import os
import posixpath
import re
import shutil
import sqlite3
import tempfile
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import BinaryIO, Iterable

from ..config import get_settings
from ..vault import paths
from .projects import now_iso, project_dir
from . import trash as trash_service

CHUNK_SIZE = 1024 * 1024  # 1MB
# 문서가 놓인 폴더(과제 폴더 기준 상대 경로). 링크는 이 위치를 기준으로 만든다.
ENTRY_DOC_DIR = "logs"
PROJECT_DOC_DIR = ""
MIN_FREE_BYTES = 200 * 1024 * 1024  # 여유 공간이 이보다 적으면 중단한다
THUMB_MAX = 480
_UNSAFE = re.compile(r"[\\/:*?\"<>|\x00-\x1f]+")
_REF_PATTERN = re.compile(r"\]\(\s*([^)\s]+)")


def safe_filename(name: str) -> str:
    """경로 구분자와 제어문자를 제거한다. 한글과 공백은 유지한다.

    윈도우에서도 그대로 쓸 수 있도록 금지 문자·예약어·길이까지 함께 정리한다.
    """
    name = unicodedata.normalize("NFC", name or "")
    name = _UNSAFE.sub("", name).strip().lstrip(".")
    name = paths.windows_safe_filename(name)
    if len(name) > paths.MAX_FILENAME_LEN:
        stem, dot, suffix = name.rpartition(".")
        if dot and len(suffix) <= 10:
            keep = paths.MAX_FILENAME_LEN - len(suffix) - 1
            name = f"{stem[:keep].rstrip(' .')}.{suffix}"
        else:
            name = name[: paths.MAX_FILENAME_LEN].rstrip(" .")
    return name or "attachment"


def is_image(mime: str | None) -> bool:
    return bool(mime and mime.startswith("image/"))


def guess_mime(name: str, fallback: str | None) -> str:
    if fallback and fallback != "application/octet-stream":
        return fallback
    return mimetypes.guess_type(name)[0] or "application/octet-stream"


def _stream_to_temp(source: BinaryIO, directory: Path) -> tuple[Path, str, int]:
    """청크로 임시 파일에 쓰면서 같은 패스에 sha256과 크기를 계산한다."""
    directory.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    size = 0
    fd, tmp_name = tempfile.mkstemp(dir=str(directory), suffix=".part")
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "wb") as target:
            while True:
                chunk = source.read(CHUNK_SIZE)
                if not chunk:
                    break
                if shutil.disk_usage(directory).free < MIN_FREE_BYTES:
                    raise OSError("디스크 여유 공간이 부족합니다.")
                target.write(chunk)
                digest.update(chunk)
                size += len(chunk)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    return tmp_path, digest.hexdigest(), size


def save_attachment(
    conn: sqlite3.Connection,
    project_id: str,
    filename: str,
    source: BinaryIO,
    bucket_rel: str,
    doc_dir: str,
    entry_id: int | None = None,
    report_id: int | None = None,
    content_type: str | None = None,
) -> dict:
    """첨부를 bucket_rel(과제 폴더 기준) 아래에 저장한다.

    doc_dir는 이 첨부를 링크할 문서가 놓인 폴더로, 마크다운 링크를 만드는 데 쓴다.
    """
    directory = project_dir(conn, project_id)
    bucket = paths.safe_join(directory, *bucket_rel.split("/"))
    tmp_path, sha256, size = _stream_to_temp(source, bucket)

    clean_name = safe_filename(filename)
    # 같은 과제에 내용도 이름도 같은 파일이 이미 있으면 다시 저장하지 않는다.
    # 내용만 같고 이름이 다르면 별개 파일로 둔다 — 빈 파일이나 같은 템플릿에서
    # 파생된 서로 다른 자료가 한 파일로 합쳐지면 이름과 확장자가 어긋난다.
    existing = conn.execute(
        "SELECT * FROM attachment WHERE project_id = ? AND sha256 = ? AND orig_name = ?",
        (project_id, sha256, clean_name),
    ).fetchone()
    if existing and (directory / existing["rel_path"]).exists():
        tmp_path.unlink(missing_ok=True)
        return dict(existing) | {"deduplicated": True, "doc_dir": doc_dir}

    target = bucket / f"{paths.next_sequence_prefix(bucket)}-{clean_name}"
    if target.exists():
        target = paths.unique_path(bucket, target.stem, target.suffix)
    os.replace(tmp_path, target)

    rel_path = target.relative_to(directory).as_posix()
    conn.execute(
        """
        INSERT INTO attachment(project_id, entry_id, report_id, rel_path, orig_name, mime,
                               size_bytes, sha256, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(project_id, rel_path) DO UPDATE SET
          entry_id=excluded.entry_id, report_id=excluded.report_id, orig_name=excluded.orig_name,
          mime=excluded.mime, size_bytes=excluded.size_bytes, sha256=excluded.sha256
        """,
        (
            project_id,
            entry_id,
            report_id,
            rel_path,
            clean_name,
            guess_mime(clean_name, content_type),
            size,
            sha256,
            now_iso(),
        ),
    )
    conn.commit()
    sync_doc_meta(conn, entry_id=entry_id, report_id=report_id)
    row = conn.execute(
        "SELECT * FROM attachment WHERE project_id = ? AND rel_path = ?", (project_id, rel_path)
    ).fetchone()
    return dict(row) | {"deduplicated": False, "doc_dir": doc_dir}


def sync_doc_meta(
    conn: sqlite3.Connection, entry_id: int | None = None, report_id: int | None = None
) -> None:
    """문서 front matter의 attachments 목록을 실제 첨부와 맞춘다.

    본문에서 참조하지 않는 자료도 md 파일만 보고 알 수 있게 하기 위함이다.
    경로는 과제 폴더 기준으로 적어 어디서 읽어도 가리키는 대상이 분명하다.
    """
    from ..vault import markdown as md

    if entry_id is not None:
        from .entries import entry_path

        try:
            _, path = entry_path(conn, entry_id)
        except KeyError:
            return
        column, owner_id = "entry_id", entry_id
    elif report_id is not None:
        from .reports import report_path

        try:
            _, path = report_path(conn, report_id)
        except KeyError:
            return
        column, owner_id = "report_id", report_id
    else:
        return

    rel_paths = [
        row["rel_path"]
        for row in conn.execute(
            f"SELECT rel_path FROM attachment WHERE {column} = ? ORDER BY rel_path", (owner_id,)
        )
    ]
    doc = md.load(path)
    if doc.meta.get("attachments") == rel_paths:
        return
    doc.meta = md.merge_meta(doc.meta, {"attachments": rel_paths})
    md.save(path, doc)
    # 도구가 스스로 쓴 변경이므로, 외부 편집 감지가 오탐하지 않도록 기준 시각을 갱신한다.
    table = "entry" if entry_id is not None else "report"
    conn.execute(
        f"UPDATE {table} SET file_mtime = ? WHERE id = ?", (path.stat().st_mtime, owner_id)
    )
    conn.commit()


def attachment_file(conn: sqlite3.Connection, attachment_id: int) -> tuple[sqlite3.Row, Path]:
    row = conn.execute("SELECT * FROM attachment WHERE id = ?", (attachment_id,)).fetchone()
    if row is None:
        raise KeyError(attachment_id)
    directory = project_dir(conn, row["project_id"])
    return row, paths.safe_join(directory, row["rel_path"])


SPREADSHEET_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
PREVIEW_MAX_ROWS = 60
PREVIEW_MAX_COLS = 12


def is_spreadsheet(mime: str | None) -> bool:
    return mime in SPREADSHEET_MIMES


def spreadsheet_preview(conn: sqlite3.Connection, attachment_id: int) -> dict:
    """엑셀 파일에서 표와 삽입 이미지를 뽑아낸다.

    서식까지 재현하지는 않는다. "그때 무엇을 보고했는지" 확인이 목적이고,
    원본이 필요하면 다운로드 링크를 쓰면 된다.
    """
    row, path = attachment_file(conn, attachment_id)
    if not is_spreadsheet(row["mime"]):
        raise ValueError("엑셀 파일이 아닙니다.")

    import base64

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for cells in worksheet.iter_rows(
            max_row=PREVIEW_MAX_ROWS, max_col=PREVIEW_MAX_COLS, values_only=True
        ):
            values = ["" if cell is None else str(cell) for cell in cells]
            if any(value.strip() for value in values):
                rows.append(values)

        images = []
        for image in getattr(worksheet, "_images", []):  # openpyxl 공개 API가 없다
            try:
                data = image._data()
                images.append("data:image/png;base64," + base64.b64encode(data).decode())
            except Exception:
                continue

        if rows or images:
            sheets.append({"name": worksheet.title, "rows": rows, "images": images})

    workbook.close()
    return {
        "orig_name": row["orig_name"],
        "sheets": sheets,
        "truncated": any(len(sheet["rows"]) >= PREVIEW_MAX_ROWS for sheet in sheets),
    }


def thumbnail(conn: sqlite3.Connection, attachment_id: int) -> Path:
    """썸네일은 vault/.index/thumbs 아래 캐시한다 (과제 폴더를 어지럽히지 않는다)."""
    row, path = attachment_file(conn, attachment_id)
    if not is_image(row["mime"]):
        raise ValueError("이미지가 아닙니다.")

    cache_dir = get_settings().index_dir / "thumbs"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_dir / f"{row['sha256']}.jpg"
    if cached.exists():
        return cached

    from PIL import Image

    with Image.open(path) as image:
        image = image.convert("RGB")
        image.thumbnail((THUMB_MAX, THUMB_MAX))
        image.save(cached, "JPEG", quality=82)
    return cached


def markdown_link(rel_path: str, orig_name: str, doc_dir: str, image: bool) -> str:
    """문서가 놓인 폴더에서 본 상대 링크를 만든다.

    logs/2026-09-03-x.md 에서 assets/… 를 가리키려면 `../assets/…` 여야
    VS Code·GitHub·옵시디언 등 외부 뷰어에서도 파일을 찾는다.
    """
    link = posixpath.relpath(rel_path, doc_dir) if doc_dir else rel_path
    return f"{'!' if image else ''}[{orig_name}]({link})"


def resolve_link(link: str, doc_dir: str) -> str | None:
    """문서 기준 상대 링크를 과제 폴더 기준 경로로 되돌린다."""
    if not link or "://" in link or link.startswith(("/", "#")):
        return None
    resolved = posixpath.normpath(posixpath.join(doc_dir, link)) if doc_dir else posixpath.normpath(link)
    return None if resolved.startswith("..") else resolved


def referenced_paths(documents: Iterable[tuple[str, str]]) -> set[str]:
    """(본문, 문서 폴더) 목록에서 참조하는 첨부 경로를 과제 폴더 기준으로 모은다."""
    found: set[str] = set()
    for body, doc_dir in documents:
        for link in _REF_PATTERN.findall(body or ""):
            resolved = resolve_link(link, doc_dir)
            if resolved:
                found.add(resolved)
    return found


def list_attachments(conn: sqlite3.Connection, project_id: str) -> list[dict]:
    """과제의 첨부 목록. 본문에서 참조되지 않는 파일은 orphan으로 표시만 한다."""
    documents: list[tuple[str, str]] = [
        (row["body"] or "", ENTRY_DOC_DIR)
        for row in conn.execute("SELECT body FROM entry WHERE project_id = ?", (project_id,))
    ]
    documents += [
        (row["body"] or "", PROJECT_DOC_DIR)
        for row in conn.execute("SELECT body FROM project WHERE id = ?", (project_id,))
    ]
    documents += [
        (row["body"] or "", posixpath.dirname(row["rel_path"]))
        for row in conn.execute("SELECT body, rel_path FROM report WHERE project_id = ?", (project_id,))
    ]
    referenced = referenced_paths(documents)
    rows = conn.execute(
        "SELECT * FROM attachment WHERE project_id = ? ORDER BY rel_path", (project_id,)
    ).fetchall()
    return [dict(row) | {"orphan": row["rel_path"] not in referenced} for row in rows]


def delete_attachment(conn: sqlite3.Connection, attachment_id: int) -> None:
    row, path = attachment_file(conn, attachment_id)
    trash = get_settings().trash_dir
    trash.mkdir(parents=True, exist_ok=True)
    if path.exists():
        target = paths.unique_path(
            trash, f"{row['project_id']}-{datetime.now():%Y%m%d%H%M%S}-{path.stem}", path.suffix
        )
        paths.move(path, target)
        trash_service.record(
            "attachment",
            label=row["orig_name"],
            moved_to=target,
            origin=path,
            project_id=row["project_id"],
        )
    conn.execute("DELETE FROM attachment WHERE id = ?", (attachment_id,))
    conn.commit()
    sync_doc_meta(conn, entry_id=row["entry_id"], report_id=row["report_id"])
